import string
import openpyxl

# 获取Excel文件的整体工作簿
gongzuobu1 = openpyxl.load_workbook(
    "python_test\初始数据\线上销售-商品\多点线上业绩_202203021040.xlsx")

# 获取工作簿1的所有工作表名称（以列表的形式呈现）
print(gongzuobu1.get_sheet_names())

# 也可以用这样的方式获取
print(gongzuobu1.sheetnames)

# 获取指定的工作表
gongzuobiao1 = gongzuobu1['Sheet0']

# 创建一个工作表（默认是在上一个工作表的后面接；但是可以通过参数修改）
# index参数修改插入位置，title参数修改名称
gongzuobiao2 = gongzuobu1.create_sheet(index=0, title='test')
print(gongzuobu1.sheetnames)

# 删除一个工作表（传入的参数应为一个工作表对象）
gongzuobu1.remove(gongzuobiao2)
print(gongzuobu1.sheetnames)

# 定位单元格
danyuange1 = gongzuobiao1['F2']
# 第几行，第几列，单元格是哪个
print(danyuange1.row, danyuange1.column, danyuange1.coordinate)
# 获取单元格内容
print(danyuange1.value)
# 找偏移单元格（往下，往右）
print(danyuange1.offset(2, 1))

# 十进制转二十六进制
print(openpyxl.utils.cell.get_column_letter(496))

# 二十六进制转十进制
print(openpyxl.utils.cell.column_index_from_string('JB'))

# 遍历单元格内容（切片范围）
for each_hang in gongzuobiao1['E1':'F20']:
    for each_lie in each_hang:
        print(each_lie.value, end='     ')
    print()

for hang in range(1, 21):
    for lie in ['E', 'F']:
        cell_name = lie + str(hang)
        print(gongzuobiao1[cell_name].value, end='    ')
    print()

# 获取第任意列
for hang in gongzuobiao1.rows:
    print(hang[0].value)

# 指定起始和终止位置()
for hang in gongzuobiao1.iter_rows(min_row=3, min_col=2, max_row=7,
                                   max_col=10):
    print(hang[0].value)

# 拷贝工作表
gongzuobiao3 = gongzuobu1.copy_worksheet(gongzuobiao1)
gongzuobu1.save("python_test\初始数据\线上销售-商品\多点线上业绩_202203021040.xlsx")
