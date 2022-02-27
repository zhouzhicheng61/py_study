import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
print(ws.title)
ws['A1'] = 520
ws.append([1, 2, 3])  #a2,b2,c2
import datetime

ws['A3'] = datetime.datetime.now()
wb.save('excel-01.xlsx')